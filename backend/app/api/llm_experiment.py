from __future__ import annotations

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import ValidationError
from typing import Any, Dict, List, Optional
import io
import json
import os
import re
import subprocess
import sys
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime
from contextlib import redirect_stdout

from app.api.endpoints import get_current_user
from app.models.user import User
from app.models.simulation import Simulation
from app.models.llm_experiment import (
    LLMExperimentAnalysisFileRecord,
    LLMExperimentAnalysisRecord,
    LLMExperimentAnalyzeRunRequest,
    LLMExperimentAnalyzeRunResponse,
    LLMExperimentBuildRequest,
    LLMExperimentBuildResponse,
    LLMExperimentComparisonReportRequest,
    LLMExperimentComparisonReportResponse,
    LLMExperimentDesignRequest,
    LLMExperimentDesignResponse,
    LLMExperimentPaperRecord,
    LLMExperimentSessionRecord,
    LLMExperimentUploadAnalysisFileResponse,
    LLMExperimentUploadPaperResponse,
)
from app.core.adapter import ai_chat
from app.core.file_utils import extract_text_from_file
from app.core.simulation_generator import generate_simulation_config
from app.core.storage import load_data, save_data


router = APIRouter(prefix="/llm-experiments", tags=["llm-experiments"])


def _as_dict_store(value: Any, *, key: Optional[str] = None) -> Dict[str, Any]:
    """Normalize persisted storage payloads to dict-shaped stores.

    Some historical data files may contain []/null due to earlier schema versions.
    This keeps runtime logic resilient without requiring manual data migration.
    """
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        if key is None:
            return {}
        return {key: value}
    return {}


_raw_llm_db: Dict[str, Any] = load_data(
    "llm_experiments.json",
    {"papers": {}, "sessions": [], "analyses": [], "analysis_files": []},
)
_raw_llm_db = _as_dict_store(_raw_llm_db)
LLM_DB: Dict[str, Any] = {
    "papers": _raw_llm_db.get("papers") if isinstance(_raw_llm_db.get("papers"), dict) else {},
    "sessions": _raw_llm_db.get("sessions") if isinstance(_raw_llm_db.get("sessions"), list) else [],
    "analyses": _raw_llm_db.get("analyses") if isinstance(_raw_llm_db.get("analyses"), list) else [],
    "analysis_files": _raw_llm_db.get("analysis_files") if isinstance(_raw_llm_db.get("analysis_files"), list) else [],
}

AGENTS_DB: Dict[str, Dict[str, Any]] = _as_dict_store(load_data("agents.json", {}))
SIMULATIONS_DB: Dict[str, Dict[str, Any]] = _as_dict_store(load_data("simulations.json", {}))


def _save_llm_db() -> None:
    save_data("llm_experiments.json", LLM_DB)


def _save_agents_db() -> None:
    save_data("agents.json", AGENTS_DB)


def _save_simulations_db() -> None:
    save_data("simulations.json", SIMULATIONS_DB)


def _extract_json_object(text: str) -> Dict[str, Any]:
    if not text:
        return {}
    raw = text.strip()
    if raw.startswith("```"):
        raw = raw.replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except Exception:
        pass
    match = re.search(r"\{[\s\S]*\}", raw)
    if not match:
        return {}
    try:
        return json.loads(match.group(0))
    except Exception:
        return {}


def _run_with_timeout(fn, timeout_s: float, *args, **kwargs):
    """Run blocking stage functions with a hard timeout.

    This prevents a single long model call from hanging the entire solve pipeline.
    """
    with ThreadPoolExecutor(max_workers=1) as executor:
        fut = executor.submit(fn, *args, **kwargs)
        return fut.result(timeout=max(1.0, float(timeout_s)))


def _to_timestamp(value: Any) -> float:
    if value is None:
        return time.time()
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return time.time()
        try:
            return float(s)
        except Exception:
            pass
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.timestamp()
        except Exception:
            return time.time()
    return time.time()


def _resolve_user_chat_config(current_user: User) -> Dict[str, str]:
    candidates = [
        a for a in AGENTS_DB.values()
        if a.get("user_id") == current_user.username
    ]
    for a in candidates:
        if a.get("provider") == "deepseek" and a.get("api_key"):
            return {
                "model": a.get("model") or os.environ.get("DEEPSEEK_DEFAULT_MODEL") or "deepseek-chat",
                "api_key": a.get("api_key") or "",
                "base_url": a.get("base_url") or os.environ.get("DEEPSEEK_BASE_URL") or "https://api.deepseek.com",
            }
    return {
        "model": os.environ.get("DEEPSEEK_DEFAULT_MODEL") or "deepseek-chat",
        "api_key": os.environ.get("DEEPSEEK_API_KEY") or "",
        "base_url": os.environ.get("DEEPSEEK_BASE_URL") or "https://api.deepseek.com",
    }


def _ai_text(
    system_prompt: str,
    user_prompt: str,
    *,
    temperature: float = 0.3,
    max_tokens: int = 4096,
    model: Optional[str] = None,
    api_key: Optional[str] = None,
    base_url: Optional[str] = None,
) -> str:
    result = ai_chat(
        messages_or_prompt=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        model=model,
        api_key=api_key,
        base_url=base_url,
        temperature=temperature,
        max_tokens=max_tokens,
        timeout_s=240.0,
    )
    if isinstance(result, dict):
        if result.get("error"):
            return ""
        content = result.get("content")
        if content:
            return str(content)
        return json.dumps(result, ensure_ascii=False)
    return str(result or "")


def _resolve_paper_text_for_user(current_user: User, paper_id: Optional[str], paper_text: Optional[str]) -> str:
    if paper_text and paper_text.strip():
        return paper_text.strip()
    if not paper_id:
        return ""
    rec = (LLM_DB.get("papers", {}) or {}).get(paper_id)
    if not rec or rec.get("user_id") != current_user.username:
        raise HTTPException(status_code=404, detail="Paper not found")
    return str(rec.get("content") or "")


def _resolve_paper_text_with_session_fallback(
    current_user: User,
    *,
    paper_id: Optional[str],
    paper_text: Optional[str],
    session_id: Optional[str],
) -> str:
    # Prefer explicit paper inputs first.
    text = _resolve_paper_text_for_user(current_user, paper_id, paper_text)
    if text.strip():
        return text

    # If caller only has session_id, recover paper_id from session.
    if session_id:
        sessions = LLM_DB.get("sessions", []) or []
        hit = next(
            (
                s
                for s in reversed(sessions)
                if s.get("id") == session_id and s.get("user_id") == current_user.username
            ),
            None,
        )
        if hit:
            sid_paper_id = hit.get("paper_id")
            return _resolve_paper_text_for_user(current_user, sid_paper_id, None)

    return ""


def _extract_analysis_file_text(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        try:
            import pandas as pd  # type: ignore

            excel = pd.ExcelFile(file_path)
            parts: List[str] = []
            for sheet in excel.sheet_names[:5]:
                df = pd.read_excel(file_path, sheet_name=sheet)
                preview = df.head(80).to_csv(index=False)
                parts.append(f"=== Sheet: {sheet} ===\n{preview}")
            return "\n\n".join(parts)[:80000]
        except Exception:
            pass
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(file_path, data_only=True)
            chunks: List[str] = []
            for ws in wb.worksheets[:5]:
                rows: List[str] = []
                for row in ws.iter_rows(min_row=1, max_row=80, values_only=True):
                    cells = ["" if c is None else str(c) for c in row]
                    rows.append(",".join(cells))
                chunks.append(f"=== Sheet: {ws.title} ===\n" + "\n".join(rows))
            return "\n\n".join(chunks)[:80000]
        except Exception:
            return f"[Uploaded analysis file: {os.path.basename(file_path)}. Could not parse xlsx content in current environment.]"
    return extract_text_from_file(file_path) or ""


def _resolve_analysis_file_text_for_user(current_user: User, analysis_file_id: Optional[str]) -> str:
    if not analysis_file_id:
        return ""
    rows = LLM_DB.get("analysis_files", []) or []
    hit = next((r for r in rows if r.get("id") == analysis_file_id and r.get("user_id") == current_user.username), None)
    if not hit:
        raise HTTPException(status_code=404, detail="Analysis file not found")
    return str(hit.get("content") or "")


def _upsert_user_agent(agent_id: str, user_id: str, persona: str) -> bool:
    if not agent_id:
        return False
    existing = AGENTS_DB.get(agent_id)
    if existing:
        return False

    template = None
    for a in AGENTS_DB.values():
        if a.get("user_id") == user_id:
            template = a
            break

    if template:
        new_agent = dict(template)
        new_agent["id"] = agent_id
        new_agent["name"] = agent_id.replace("-", " ").replace("_", " ").title()
        new_agent["persona"] = persona
        new_agent["long_term_memory"] = []
        new_agent["user_id"] = user_id
        AGENTS_DB[agent_id] = new_agent
        return True

    AGENTS_DB[agent_id] = {
        "id": agent_id,
        "name": agent_id.replace("-", " ").replace("_", " ").title(),
        "provider": "deepseek",
        "model": os.environ.get("DEEPSEEK_DEFAULT_MODEL", "deepseek-chat"),
        "base_url": os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com"),
        "api_key": os.environ.get("DEEPSEEK_API_KEY"),
        "persona": persona,
        "long_term_memory": [],
        "relationships": [],
        "files": [],
        "usage_example": None,
        "user_id": user_id,
    }
    return True


def _collect_agent_ids(simulation: Dict[str, Any]) -> List[str]:
    output: List[str] = []

    def _walk(steps: List[Dict[str, Any]]) -> None:
        for step in steps or []:
            if not isinstance(step, dict):
                continue
            ids = step.get("agent_ids")
            if isinstance(ids, list):
                for aid in ids:
                    if isinstance(aid, str) and aid.strip() and aid not in output:
                        output.append(aid.strip())
            if step.get("inner_steps") and isinstance(step.get("inner_steps"), list):
                _walk(step.get("inner_steps"))

    _walk(simulation.get("steps") or [])
    return output


def _has_steps(simulation: Dict[str, Any]) -> bool:
    return isinstance(simulation, dict) and isinstance(simulation.get("steps"), list) and len(simulation.get("steps") or []) > 0


def _persist_simulation_for_user(simulation: Dict[str, Any], current_user: User) -> None:
    sim_id = str(uuid.uuid4())
    rec = {
        "id": sim_id,
        "user_id": current_user.username,
        "name": simulation.get("name") or "LLM Experiment Simulation",
        "description": simulation.get("description") or "Generated by LLM experiment pipeline",
        "variables": simulation.get("variables") or [],
        "steps": simulation.get("steps") or [],
        "created_at": time.time(),
    }
    SIMULATIONS_DB[sim_id] = rec
    _save_simulations_db()


def _default_analysis_code() -> str:
    return (
        "from collections import Counter\n"
        "analysis_result = {}\n"
        "analysis_result['history_count'] = len(run_history)\n"
        "analysis_result['state_keys'] = sorted(list((final_world_state or {}).keys()))\n"
        "analysis_result['paper_excerpt_len'] = len(paper_text or '')\n"
        "analysis_result['analysis_data_len'] = len(analysis_data_text or '')\n"
        "analysis_result['analysis_injected_requirements'] = analysis_injected_requirements or ''\n"
        "agent_counter = Counter()\n"
        "for item in run_history:\n"
        "    name = item.get('agent_name') or 'Unknown'\n"
        "    agent_counter[name] += 1\n"
        "analysis_result['agent_message_count'] = dict(agent_counter)\n"
        "numeric_state = {}\n"
        "for k, v in (final_world_state or {}).items():\n"
        "    if isinstance(v, (int, float)):\n"
        "        numeric_state[k] = v\n"
        "analysis_result['numeric_state'] = numeric_state\n"
        "print('Analysis summary ready.')\n"
    )


def _run_analysis_code(
    code: str,
    run_history: List[Dict[str, Any]],
    final_world_state: Dict[str, Any],
    requirements: str,
    experiment_goal: Optional[str],
    paper_text: str,
    analysis_data_text: str,
    analysis_injected_requirements: str,
) -> tuple[str, Dict[str, Any]]:
    safe_globals: Dict[str, Any] = {"__builtins__": __builtins__}
    safe_locals: Dict[str, Any] = {
        "run_history": run_history,
        "final_world_state": final_world_state,
        "requirements": requirements,
        "experiment_goal": experiment_goal,
        "paper_text": paper_text,
        "analysis_data_text": analysis_data_text,
        "analysis_injected_requirements": analysis_injected_requirements,
        "analysis_result": {},
    }

    def _normalize_module_name(name: str) -> str:
        if not name:
            return ""
        n = str(name).strip().strip("'\"")
        if not n:
            return ""
        if n in {"cv2"}:
            return "opencv-python-headless"
        if n in {"sklearn"}:
            return "scikit-learn"
        if n in {"PIL"}:
            return "pillow"
        return n.split(".")[0]

    def _extract_missing_module(exc: Exception) -> str:
        mod = _normalize_module_name(getattr(exc, "name", ""))
        if mod:
            return mod
        m = re.search(r"No module named ['\"]([^'\"]+)['\"]", str(exc))
        if m:
            return _normalize_module_name(m.group(1))
        return ""

    def _install_module(module_name: str) -> tuple[bool, str]:
        if not module_name:
            return False, "empty module name"
        try:
            cmd = [sys.executable, "-m", "pip", "install", module_name]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
            ok = proc.returncode == 0
            log = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
            return ok, log.strip()
        except Exception as e:
            return False, str(e)

    def _exec_once() -> tuple[str, Dict[str, Any]]:
        output = io.StringIO()
        with redirect_stdout(output):
            exec(code, safe_globals, safe_locals)
        result = safe_locals.get("analysis_result")
        if not isinstance(result, dict):
            result = {"value": result}
        return output.getvalue(), result

    tried_modules = set()
    while True:
        try:
            return _exec_once()
        except (ModuleNotFoundError, ImportError) as ie:
            missing_mod = _extract_missing_module(ie)
            if not missing_mod or missing_mod in tried_modules:
                raise
            tried_modules.add(missing_mod)
            ok, install_log = _install_module(missing_mod)
            if not ok:
                raise RuntimeError(f"Auto-install failed for '{missing_mod}': {install_log}")
        except Exception:
            raise


def _build_stage1_design(
    paper_text: str,
    req: LLMExperimentBuildRequest,
    chat_cfg: Optional[Dict[str, str]] = None,
) -> tuple[str, str]:
    system_prompt = (
        "你是模拟创作AI。你会基于论文文本和用户要求，完整设计一个可复现实验。"
        "输出必须可执行、步骤完整、指标可量化。"
        "另外请给出后续分析思路（分析目标、指标、方法、可视化建议）。"
    )
    user_prompt = (
        f"实验目标: {req.experiment_goal or '未指定'}\n"
        f"约束条件: {req.constraints or '无'}\n"
        f"用户要求: {req.requirements}\n\n"
        f"论文内容:\n{paper_text[:40000]}\n\n"
        "请严格输出 JSON: {\"experiment_design\":\"...\",\"analysis_thinking\":\"...\"}"
    )
    raw = _ai_text(
        system_prompt,
        user_prompt,
        temperature=0.2,
        max_tokens=4096,
        model=(chat_cfg or {}).get("model"),
        api_key=(chat_cfg or {}).get("api_key"),
        base_url=(chat_cfg or {}).get("base_url"),
    )
    obj = _extract_json_object(raw)

    design = str(obj.get("experiment_design") or "")
    thinking = str(obj.get("analysis_thinking") or "")

    if not design:
        design = (
            "1) 明确实验对象与基准场景；2) 定义可控变量与观测窗口；"
            "3) 设计对照组与干预组；4) 规定每步决策输入输出；"
            "5) 执行多轮模拟并记录关键变量时间序列；6) 汇总统计并复核。"
        )
    if not thinking:
        thinking = (
            "围绕库存波动、缺货率、订单放大倍数、收敛速度等指标进行分析，"
            "先做描述统计，再做组间对比与关键因子敏感性分析。"
        )

    return design, thinking


def _build_stage2_simulation(
    design: str,
    paper_text: str,
    req: LLMExperimentBuildRequest,
    *,
    chat_cfg: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    sim_prompt = (
        "你是模拟创作Agent组。你负责把实验方案转换为本平台 simulation JSON。"
        "只负责模拟配置，不负责实验结果分析。\n\n"
        f"实验方案:\n{design}\n\n"
        f"用户要求:\n{req.requirements}\n\n"
        f"实验目标:{req.experiment_goal or '未指定'}\n"
        f"约束:{req.constraints or '无'}\n\n"
        "要求：steps可被 /api/simulation/run_step 逐步执行；变量请完整初始化。"
    )
    generated = generate_simulation_config(
        sim_prompt,
        file_content=paper_text[:40000],
        file_names=[],
        model=(chat_cfg or {}).get("model"),
        api_key=(chat_cfg or {}).get("api_key"),
        base_url=(chat_cfg or {}).get("base_url"),
    )
    if not isinstance(generated, dict):
        raise HTTPException(status_code=500, detail="Simulation generation failed")
    return generated


def _build_stage2_fallback_simulation(
    *,
    req: LLMExperimentBuildRequest,
    design: str,
    reason: str,
) -> Dict[str, Any]:
    """Build a deterministic, executable simulation when AI generation is unavailable."""
    sim_id = f"llm-fallback-{uuid.uuid4().hex[:8]}"
    req_text = (req.requirements or "").strip()
    goal_text = (req.experiment_goal or "").strip()
    constraint_text = (req.constraints or "").strip()
    design_text = (design or "").strip()[:800]

    return {
        "id": sim_id,
        "name": "LLM Fallback Simulation: Chapter-5 Human vs Agent",
        "description": (
            "自动降级模板：用于在 AI 适配器不可用时，仍可执行‘人类实验基线 vs Agent替代’比较流程。"
        ),
        "variables": [
            {"key": "paper_focus", "value": json.dumps("第五章人类实验", ensure_ascii=False), "description": "论文焦点章节"},
            {"key": "requirements_text", "value": json.dumps(req_text, ensure_ascii=False), "description": "用户实验要求"},
            {"key": "goal_text", "value": json.dumps(goal_text, ensure_ascii=False), "description": "实验目标"},
            {"key": "constraint_text", "value": json.dumps(constraint_text, ensure_ascii=False), "description": "实验约束"},
            {"key": "design_outline", "value": json.dumps(design_text, ensure_ascii=False), "description": "实验设计摘要"},
            {"key": "human_reference", "value": "{}", "description": "原文人类实验基线"},
            {"key": "agent_result", "value": "{}", "description": "Agent替代实验结果"},
            {"key": "comparison_summary", "value": json.dumps("", ensure_ascii=False), "description": "比较结论"},
            {"key": "fallback_reason", "value": json.dumps(reason[:500], ensure_ascii=False), "description": "触发降级的原因"},
        ],
        "steps": [
            {
                "id": "step-init-human-baseline",
                "type": "code",
                "code_snippet": (
                    "state['human_reference'] = {\n"
                    "  'chapter': state.get('paper_focus', '第五章'),\n"
                    "  'note': '请在分析阶段结合原文第五章结果校正该基线数值',\n"
                    "  'kpi': {'service_level': 0.82, 'stockout_rate': 0.18, 'cost_index': 1.00}\n"
                    "}\n"
                    "print('Initialized human baseline from template.')"
                ),
                "output_var": "human_reference",
                "execution_mode": "serial",
                "use_rag": False,
            },
            {
                "id": "step-run-agent-replacement",
                "type": "code",
                "code_snippet": (
                    "h = state.get('human_reference', {}).get('kpi', {})\n"
                    "service = float(h.get('service_level', 0.82)) + 0.03\n"
                    "stockout = max(0.0, float(h.get('stockout_rate', 0.18)) - 0.03)\n"
                    "cost = float(h.get('cost_index', 1.0)) + 0.04\n"
                    "state['agent_result'] = {\n"
                    "  'method': 'agent_replacement',\n"
                    "  'kpi': {'service_level': round(service, 3), 'stockout_rate': round(stockout, 3), 'cost_index': round(cost, 3)}\n"
                    "}\n"
                    "print('Generated fallback agent replacement result.')"
                ),
                "output_var": "agent_result",
                "execution_mode": "serial",
                "use_rag": False,
            },
            {
                "id": "step-compare-human-vs-agent",
                "type": "code",
                "code_snippet": (
                    "h = state.get('human_reference', {}).get('kpi', {})\n"
                    "a = state.get('agent_result', {}).get('kpi', {})\n"
                    "delta_service = round(float(a.get('service_level', 0)) - float(h.get('service_level', 0)), 3)\n"
                    "delta_stockout = round(float(a.get('stockout_rate', 0)) - float(h.get('stockout_rate', 0)), 3)\n"
                    "delta_cost = round(float(a.get('cost_index', 0)) - float(h.get('cost_index', 0)), 3)\n"
                    "state['comparison_summary'] = {\n"
                    "  'task': '第五章: 人类实验 vs Agent替代实验',\n"
                    "  'delta': {'service_level': delta_service, 'stockout_rate': delta_stockout, 'cost_index': delta_cost},\n"
                    "  'comment': '这是后端降级模板生成的可执行结果，请在分析阶段结合原文与CSV进一步校准。'\n"
                    "}\n"
                    "print('Comparison summary prepared.')"
                ),
                "output_var": "comparison_summary",
                "execution_mode": "serial",
                "use_rag": False,
            },
        ],
    }


def _build_stage3_review(
    simulation_draft: Dict[str, Any],
    paper_text: str,
    req: LLMExperimentBuildRequest,
    chat_cfg: Optional[Dict[str, str]] = None,
) -> tuple[Dict[str, Any], str]:
    system_prompt = (
        "你是检查AI。你要阅读论文和用户要求，并理解当前平台 simulation 运行逻辑。"
        "请改进模拟配置，使它更稳定、可执行、且更贴合论文方法。"
    )
    user_prompt = (
        f"用户要求: {req.requirements}\n"
        f"实验目标: {req.experiment_goal or '未指定'}\n"
        f"约束: {req.constraints or '无'}\n\n"
        f"论文内容:\n{paper_text[:24000]}\n\n"
        f"当前 simulation 草稿:\n{json.dumps(simulation_draft, ensure_ascii=False)[:30000]}\n\n"
        "请严格输出 JSON: {\"improved_simulation\":{...},\"checker_notes\":\"...\"}"
    )

    raw = _ai_text(
        system_prompt,
        user_prompt,
        temperature=0.2,
        max_tokens=6144,
        model=(chat_cfg or {}).get("model"),
        api_key=(chat_cfg or {}).get("api_key"),
        base_url=(chat_cfg or {}).get("base_url"),
    )
    obj = _extract_json_object(raw)

    improved = obj.get("improved_simulation") if isinstance(obj, dict) else None
    notes = str(obj.get("checker_notes") or "") if isinstance(obj, dict) else ""

    if not isinstance(improved, dict):
        improved = simulation_draft
    if not notes:
        notes = "检查AI未返回结构化说明，沿用草稿或最小修正版本。"

    return improved, notes


def _seed_analysis_code(
    analysis_thinking: str,
    *,
    paper_text: str = "",
    analysis_data_text: str = "",
    analysis_injected_requirements: str = "",
    chat_cfg: Optional[Dict[str, str]] = None,
) -> str:
    if not analysis_thinking.strip():
        return _default_analysis_code()

    system_prompt = (
        "你是实验后续分析模块。请根据分析思路生成 Python 分析代码。"
        "代码输入变量: run_history(list), final_world_state(dict), requirements(str), experiment_goal(str|None), "
        "paper_text(str), analysis_data_text(str), analysis_injected_requirements(str)。"
        "必须写入 analysis_result(dict)，并可使用 print 输出关键结论。"
        "仅输出 Python 代码，不要 markdown。"
    )
    user_prompt = (
        f"分析思路:\n{analysis_thinking[:12000]}\n\n"
        f"用户注入要求:\n{analysis_injected_requirements[:4000]}\n\n"
        f"论文原文片段:\n{paper_text[:12000]}\n\n"
        f"模拟导出数据片段(csv/txt/json):\n{analysis_data_text[:12000]}"
    )
    code = _ai_text(
        system_prompt,
        user_prompt,
        temperature=0.1,
        max_tokens=2048,
        model=(chat_cfg or {}).get("model"),
        api_key=(chat_cfg or {}).get("api_key"),
        base_url=(chat_cfg or {}).get("base_url"),
    ).strip()
    if code.startswith("```"):
        code = code.replace("```python", "").replace("```", "").strip()
    if code.startswith('{') and '"error"' in code:
        return _default_analysis_code()
    if not code:
        code = _default_analysis_code()
    return code


def _fallback_comparison_report(
    requirements: str,
    compare_requirements: str,
    paper_text: str,
    analysis_result: Dict[str, Any],
    analysis_conclusion: str,
) -> str:
    has_paper = bool((paper_text or "").strip())
    result_keys = sorted(list((analysis_result or {}).keys()))
    key_preview = ", ".join(result_keys[:12]) if result_keys else "无结构化结果键"
    conclusion_preview = (analysis_conclusion or "").strip()
    if len(conclusion_preview) > 300:
        conclusion_preview = conclusion_preview[:300] + "..."

    confidence = "中"
    if not has_paper or not analysis_result:
        confidence = "低"
    elif has_paper and analysis_result:
        confidence = "中-高"

    lines = [
        "# 比较报告（系统兜底生成）",
        "",
        "## 1) 论文主张与实验设定对齐情况",
        f"- 用户主要求：{requirements or '未提供'}",
        f"- 用户比较要求：{compare_requirements or '未提供'}",
        f"- 论文文本可用：{'是' if has_paper else '否'}",
        "",
        "## 2) 实验结果与论文结论的一致/差异",
        f"- 分析结果键概览：{key_preview}",
        f"- 分析结论摘要：{conclusion_preview or '无'}",
        "- 初步判断：请重点核对关键指标的方向（上升/下降、高于/低于）是否与论文主张一致。",
        "",
        "## 3) 差异原因推测",
        "- 可能由实验参数缩减、样本轮次不足、代理策略与论文假设不完全一致导致。",
        "- 若使用了自动修复或默认分析代码，结论解释力会下降。",
        "",
        "## 4) 可信度评估与改进建议",
        f"- 当前可信度：{confidence}",
        "- 建议补充：固定随机种子、多次重复运行、报告均值与方差、保留关键中间变量。",
        "",
        "## 5) 后续实验建议（可执行）",
        "- 增加对照组（原始参数 vs 修正参数）并做同口径指标比较。",
        "- 将核心结论转换为可检验假设，并增加显著性检验/置信区间。",
    ]
    return "\n".join(lines)


@router.post("/papers/upload", response_model=LLMExperimentUploadPaperResponse)
async def upload_paper(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    upload_dir = "uploads/llm_experiments"
    os.makedirs(upload_dir, exist_ok=True)

    file_name = os.path.basename(file.filename or "paper.txt")
    path = os.path.join(upload_dir, f"{uuid.uuid4()}_{file_name}")
    with open(path, "wb") as out:
        out.write(await file.read())

    text = extract_text_from_file(path) or ""
    extracted_ok = bool(text.strip())
    if not extracted_ok:
        text = (
            f"[Paper text extraction fallback] File '{file_name}' was uploaded but text extraction returned empty. "
            "Please provide key abstract/method snippets in requirements if this is an image-scanned PDF."
        )

    paper_id = str(uuid.uuid4())
    created_at = time.time()
    rec = {
        "id": paper_id,
        "user_id": current_user.username,
        "title": file_name,
        "source": "upload",
        "file_name": file_name,
        "content": text,
        "content_preview": text[:300],
        "extracted_ok": extracted_ok,
        "created_at": created_at,
    }
    LLM_DB.setdefault("papers", {})[paper_id] = rec
    _save_llm_db()

    return LLMExperimentUploadPaperResponse(
        paper=LLMExperimentPaperRecord(
            id=paper_id,
            title=file_name,
            source="upload",
            file_name=file_name,
            content_preview=text[:300],
            created_at=created_at,
        )
    )


@router.post("/analysis-files/upload", response_model=LLMExperimentUploadAnalysisFileResponse)
async def upload_analysis_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    upload_dir = "uploads/llm_experiments/analysis_data"
    os.makedirs(upload_dir, exist_ok=True)

    file_name = os.path.basename(file.filename or "analysis_data.csv")
    path = os.path.join(upload_dir, f"{uuid.uuid4()}_{file_name}")
    with open(path, "wb") as out:
        out.write(await file.read())

    text = _extract_analysis_file_text(path)
    if not text.strip():
        text = f"[Uploaded analysis file: {file_name}]"

    file_id = str(uuid.uuid4())
    created_at = time.time()
    rec = {
        "id": file_id,
        "user_id": current_user.username,
        "file_name": file_name,
        "path": path,
        "content": text,
        "created_at": created_at,
    }
    LLM_DB.setdefault("analysis_files", []).append(rec)
    if len(LLM_DB["analysis_files"]) > 500:
        LLM_DB["analysis_files"] = LLM_DB["analysis_files"][-500:]
    _save_llm_db()

    return LLMExperimentUploadAnalysisFileResponse(
        file=LLMExperimentAnalysisFileRecord(
            id=file_id,
            file_name=file_name,
            extracted_preview=text[:400],
            created_at=created_at,
        )
    )


@router.get("/papers", response_model=List[LLMExperimentPaperRecord])
async def list_papers(current_user: User = Depends(get_current_user)):
    out: List[LLMExperimentPaperRecord] = []
    for p in (LLM_DB.get("papers", {}) or {}).values():
        if p.get("user_id") != current_user.username:
            continue
        out.append(
            LLMExperimentPaperRecord(
                id=p.get("id", ""),
                title=p.get("title", "Untitled"),
                source=p.get("source", "upload"),
                file_name=p.get("file_name"),
                content_preview=p.get("content_preview"),
                created_at=_to_timestamp(p.get("created_at")),
            )
        )
    out.sort(key=lambda x: x.created_at, reverse=True)
    return out


@router.post("/design", response_model=LLMExperimentDesignResponse)
async def build_design(req: LLMExperimentDesignRequest, current_user: User = Depends(get_current_user)):
    paper_text = _resolve_paper_text_for_user(current_user, req.paper_id, req.paper_text)
    chat_cfg = _resolve_user_chat_config(current_user)
    pseudo_req = LLMExperimentBuildRequest(
        paper_id=req.paper_id,
        paper_text=paper_text,
        requirements=req.requirements,
        experiment_goal=req.experiment_goal,
        constraints=req.constraints,
        save_simulation=False,
    )
    design, thinking = _build_stage1_design(paper_text, pseudo_req, chat_cfg=chat_cfg)
    return LLMExperimentDesignResponse(experiment_design=design, analysis_thinking=thinking)


@router.post("/cluster/solve", response_model=LLMExperimentBuildResponse)
async def solve_pipeline(req: LLMExperimentBuildRequest, current_user: User = Depends(get_current_user)):
    paper_text = _resolve_paper_text_for_user(current_user, req.paper_id, req.paper_text)
    chat_cfg = _resolve_user_chat_config(current_user)
    if not paper_text.strip():
        raise HTTPException(status_code=400, detail="Please provide paper_id or paper_text")

    design, thinking = _build_stage1_design(paper_text, req, chat_cfg=chat_cfg)
    stage2_fallback_note = ""
    stage2_timeout_s = float(os.environ.get("LLM_EXPERIMENT_STAGE2_TIMEOUT_S", "120"))
    stage3_timeout_s = float(os.environ.get("LLM_EXPERIMENT_STAGE3_TIMEOUT_S", "90"))

    try:
        simulation_draft = _run_with_timeout(
            _build_stage2_simulation,
            stage2_timeout_s,
            design,
            paper_text,
            req,
            chat_cfg=chat_cfg,
        )
    except FuturesTimeoutError:
        reason = f"stage2 timeout after {int(stage2_timeout_s)}s"
        simulation_draft = _build_stage2_fallback_simulation(req=req, design=design, reason=reason)
        stage2_fallback_note = (
            "[Fallback] Stage-2 simulation generation timeout; switched to executable template. "
            f"reason={reason}"
        )
    except Exception as e:
        reason = str(e) or e.__class__.__name__
        simulation_draft = _build_stage2_fallback_simulation(req=req, design=design, reason=reason)
        stage2_fallback_note = (
            "[Fallback] Stage-2 simulation generation failed and switched to executable template. "
            f"reason={reason}"
        )

    if not _has_steps(simulation_draft):
        simulation_draft = _build_stage2_fallback_simulation(
            req=req,
            design=design,
            reason="stage2 returned empty steps",
        )
        stage2_fallback_note = (
            "[Fallback] Stage-2 simulation generation returned empty steps and switched to executable template."
        )

    try:
        simulation_reviewed, checker_notes = _run_with_timeout(
            _build_stage3_review,
            stage3_timeout_s,
            simulation_draft,
            paper_text,
            req,
            chat_cfg=chat_cfg,
        )
    except FuturesTimeoutError:
        simulation_reviewed, checker_notes = (
            simulation_draft,
            f"[Fallback] Checker stage timeout after {int(stage3_timeout_s)}s; using simulation draft.",
        )
    except Exception as e:
        simulation_reviewed, checker_notes = (
            simulation_draft,
            f"[Fallback] Checker stage failed ({e}); using simulation draft.",
        )

    if not _has_steps(simulation_reviewed):
        simulation_reviewed = simulation_draft
        checker_notes = (checker_notes + "\n[Fallback] Checker output had empty steps; reverted to simulation draft.").strip()

    if stage2_fallback_note:
        checker_notes = (stage2_fallback_note + "\n" + (checker_notes or "")).strip()

    try:
        validated = Simulation(**{
            "id": simulation_reviewed.get("id"),
            "name": simulation_reviewed.get("name") or "LLM Experiment Simulation",
            "description": simulation_reviewed.get("description") or "",
            "steps": simulation_reviewed.get("steps") or [],
            "variables": simulation_reviewed.get("variables") or [],
        })
        simulation_reviewed = validated.dict()
    except ValidationError:
        simulation_reviewed = simulation_draft

    created_agents: List[str] = []
    for aid in _collect_agent_ids(simulation_reviewed):
        ok = _upsert_user_agent(aid, current_user.username, persona=f"You are {aid}. Follow simulation role instructions strictly.")
        if ok:
            created_agents.append(aid)
    if created_agents:
        _save_agents_db()

    if req.save_simulation:
        _persist_simulation_for_user(simulation_reviewed, current_user)

    analysis_seed = _seed_analysis_code(thinking, chat_cfg=chat_cfg)

    session_id = str(uuid.uuid4())
    created_at = time.time()
    session = LLMExperimentSessionRecord(
        id=session_id,
        user_id=current_user.username,
        paper_id=req.paper_id,
        paper_title=((LLM_DB.get("papers", {}) or {}).get(req.paper_id, {}) or {}).get("title") if req.paper_id else None,
        requirements=req.requirements,
        experiment_goal=req.experiment_goal,
        experiment_design=design,
        analysis_thinking=thinking,
        simulation_reviewed=simulation_reviewed,
        analysis_code_seed=analysis_seed,
        checker_notes=checker_notes,
        created_agent_ids=created_agents,
        created_at=created_at,
    )
    LLM_DB.setdefault("sessions", []).append(session.dict())
    if len(LLM_DB["sessions"]) > 500:
        LLM_DB["sessions"] = LLM_DB["sessions"][-500:]
    _save_llm_db()

    return LLMExperimentBuildResponse(
        session_id=session_id,
        paper_id=req.paper_id,
        experiment_design=design,
        analysis_thinking=thinking,
        simulation_draft=simulation_draft,
        simulation_reviewed=simulation_reviewed,
        checker_notes=checker_notes,
        analysis_code_seed=analysis_seed,
        created_agent_ids=created_agents,
        created_at=created_at,
    )


@router.post("/cluster/analyze-run", response_model=LLMExperimentAnalyzeRunResponse)
async def analyze_run(req: LLMExperimentAnalyzeRunRequest, current_user: User = Depends(get_current_user)):
    chat_cfg = _resolve_user_chat_config(current_user)
    analysis_thinking = req.analysis_thinking or ""
    injected_requirements = (req.analysis_injected_requirements or "").strip()
    paper_text = _resolve_paper_text_with_session_fallback(
        current_user,
        paper_id=req.paper_id,
        paper_text=None,
        session_id=req.session_id,
    )
    analysis_data_text = _resolve_analysis_file_text_for_user(current_user, req.analysis_file_id)

    if req.session_id:
        sessions = LLM_DB.get("sessions", []) or []
        hit = next((s for s in reversed(sessions) if s.get("id") == req.session_id and s.get("user_id") == current_user.username), None)
        if hit and not analysis_thinking:
            analysis_thinking = str(hit.get("analysis_thinking") or "")

    analysis_code = _seed_analysis_code(
        analysis_thinking,
        paper_text=paper_text,
        analysis_data_text=analysis_data_text,
        analysis_injected_requirements=injected_requirements,
        chat_cfg=chat_cfg,
    )
    try:
        stdout, result = _run_analysis_code(
            analysis_code,
            req.run_history or [],
            req.final_world_state or {},
            req.requirements,
            req.experiment_goal,
            paper_text,
            analysis_data_text,
            injected_requirements,
        )
    except Exception:
        analysis_code = _default_analysis_code()
        stdout, result = _run_analysis_code(
            analysis_code,
            req.run_history or [],
            req.final_world_state or {},
            req.requirements,
            req.experiment_goal,
            paper_text,
            analysis_data_text,
            injected_requirements,
        )

    conclusion_prompt = (
        "你是实验分析助手。请根据分析结果生成简洁结论，包含：关键发现、与要求的匹配度、后续建议。\n"
        f"用户要求: {req.requirements}\n"
        f"分析阶段用户注入要求: {injected_requirements or '无'}\n"
        f"实验目标: {req.experiment_goal or '未指定'}\n"
        f"论文片段: {paper_text[:4000]}\n"
        f"模拟导出数据片段: {analysis_data_text[:4000]}\n"
        f"分析结果: {json.dumps(result, ensure_ascii=False)[:12000]}\n"
        f"stdout: {stdout[:8000]}\n"
        "请输出不超过8行。"
    )
    conclusion = _ai_text(
        "你擅长从仿真实验中提炼可执行结论。",
        conclusion_prompt,
        temperature=0.2,
        max_tokens=512,
        model=chat_cfg.get("model"),
        api_key=chat_cfg.get("api_key"),
        base_url=chat_cfg.get("base_url"),
    ).strip()
    if not conclusion:
        conclusion = "分析已完成。建议优先关注关键状态变量的变化趋势与组间差异。"

    record = LLMExperimentAnalysisRecord(
        id=str(uuid.uuid4()),
        user_id=current_user.username,
        session_id=req.session_id,
        paper_id=req.paper_id,
        requirements=req.requirements,
        analysis_code=analysis_code,
        analysis_stdout=stdout,
        analysis_result=result,
        conclusion=conclusion,
        created_at=time.time(),
    )
    LLM_DB.setdefault("analyses", []).append(record.dict())
    if len(LLM_DB["analyses"]) > 800:
        LLM_DB["analyses"] = LLM_DB["analyses"][-800:]

    if req.session_id:
        sessions = LLM_DB.get("sessions", []) or []
        for s in reversed(sessions):
            if s.get("id") == req.session_id and s.get("user_id") == current_user.username:
                s["latest_analysis"] = {
                    "analysis_code": analysis_code,
                    "analysis_stdout": stdout,
                    "analysis_result": result,
                    "conclusion": conclusion,
                    "created_at": time.time(),
                }
                s["latest_analysis_file_id"] = req.analysis_file_id
                s["latest_analysis_injected_requirements"] = injected_requirements
                break
    _save_llm_db()

    return LLMExperimentAnalyzeRunResponse(
        analysis_code=analysis_code,
        analysis_stdout=stdout,
        analysis_result=result,
        conclusion=conclusion,
    )


@router.post("/cluster/comparison-report", response_model=LLMExperimentComparisonReportResponse)
async def generate_comparison_report(req: LLMExperimentComparisonReportRequest, current_user: User = Depends(get_current_user)):
    chat_cfg = _resolve_user_chat_config(current_user)
    paper_text = _resolve_paper_text_with_session_fallback(
        current_user,
        paper_id=req.paper_id,
        paper_text=None,
        session_id=req.session_id,
    )

    analysis_result = req.analysis_result
    analysis_conclusion = req.analysis_conclusion or ""
    if (not analysis_result) and req.session_id:
        analyses = LLM_DB.get("analyses", []) or []
        hit = next(
            (
                a
                for a in reversed(analyses)
                if a.get("session_id") == req.session_id and a.get("user_id") == current_user.username
            ),
            None,
        )
        if hit:
            if not analysis_result:
                analysis_result = hit.get("analysis_result") or {}
            if not analysis_conclusion:
                analysis_conclusion = str(hit.get("conclusion") or "")

    compare_prompt = (
        "你是比较报告AI。请阅读论文原文、实验分析结果与结论，输出一份比较报告。\n"
        "报告结构：\n"
        "1) 论文主张与实验设定对齐情况\n"
        "2) 实验结果与论文结论的一致/差异\n"
        "3) 差异原因推测\n"
        "4) 可信度评估与改进建议\n"
        "5) 后续实验建议（可执行）\n\n"
        f"用户主要求: {req.requirements}\n"
        f"用户比较要求: {req.compare_requirements or '无'}\n"
        f"论文片段:\n{paper_text[:40000]}\n\n"
        f"分析结果:\n{json.dumps(analysis_result or {}, ensure_ascii=False)[:14000]}\n\n"
        f"分析结论:\n{analysis_conclusion[:6000]}\n"
    )
    report = _ai_text(
        "你擅长写严谨、结构化的实验比较报告。",
        compare_prompt,
        temperature=0.2,
        max_tokens=1800,
        model=chat_cfg.get("model"),
        api_key=chat_cfg.get("api_key"),
        base_url=chat_cfg.get("base_url"),
    ).strip()
    if not report:
        report = _fallback_comparison_report(
            requirements=req.requirements,
            compare_requirements=req.compare_requirements or "",
            paper_text=paper_text,
            analysis_result=analysis_result or {},
            analysis_conclusion=analysis_conclusion,
        )

    if req.session_id:
        sessions = LLM_DB.get("sessions", []) or []
        for s in reversed(sessions):
            if s.get("id") == req.session_id and s.get("user_id") == current_user.username:
                s["latest_comparison_report"] = report
                s["latest_compare_requirements"] = req.compare_requirements or ""
                break
        _save_llm_db()

    return LLMExperimentComparisonReportResponse(report=report)


@router.get("/sessions", response_model=List[LLMExperimentSessionRecord])
async def list_sessions(limit: int = 20, current_user: User = Depends(get_current_user)):
    rows = [
        s
        for s in (LLM_DB.get("sessions", []) or [])
        if s.get("user_id") == current_user.username
    ]
    analyses = [
        a
        for a in (LLM_DB.get("analyses", []) or [])
        if a.get("user_id") == current_user.username
    ]

    latest_analysis_by_session: Dict[str, Dict[str, Any]] = {}
    for a in analyses:
        sid = str(a.get("session_id") or "").strip()
        if not sid:
            continue
        ts = _to_timestamp(a.get("created_at"))
        cur = latest_analysis_by_session.get(sid)
        cur_ts = _to_timestamp((cur or {}).get("created_at")) if cur else -1.0
        if ts >= cur_ts:
            latest_analysis_by_session[sid] = {
                "analysis_code": a.get("analysis_code") or "",
                "analysis_stdout": a.get("analysis_stdout") or "",
                "analysis_result": a.get("analysis_result") or {},
                "conclusion": a.get("conclusion") or "",
                "created_at": ts,
            }

    for row in rows:
        row["created_at"] = _to_timestamp(row.get("created_at"))
        if not row.get("paper_title") and row.get("paper_id"):
            pid = row.get("paper_id")
            row["paper_title"] = ((LLM_DB.get("papers", {}) or {}).get(pid, {}) or {}).get("title")
        if not row.get("latest_analysis"):
            sid = str(row.get("id") or "")
            if sid and sid in latest_analysis_by_session:
                row["latest_analysis"] = latest_analysis_by_session[sid]
    rows = rows[-max(1, min(limit, 100)):]
    rows.reverse()
    return [LLMExperimentSessionRecord(**r) for r in rows]
