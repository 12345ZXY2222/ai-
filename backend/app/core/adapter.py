"""Adapter module providing zhipu_chat / zhipu_embedding compatibility.

Behavior:
- If environment variable AI_PROVIDER is set to an OpenAI-compatible provider
    (currently: 'deepseek' or 'dashscope'), or the corresponding API key env var
    exists, route chat/embeddings to that provider.
- Otherwise, try to import an existing legacy zhipu adapter (if present) and call it.

Notes:
- This repo historically used the `openai` Python SDK for OpenAI-compatible APIs.
    Some environments do not have the SDK installed even if it's listed in
    requirements. We therefore provide a requests-based fallback to keep core
    functionality working.

Compatibility:
- Exposes zhipu_chat(prompt, model=..., api_key=..., temperature=..., max_tokens=...)
    which returns a string (same as legacy). Also exposes ai_chat(...) which
    returns a dict {content, reasoning_content, raw} for front-end use.
"""
from __future__ import annotations
import os
import json
from typing import Any, Dict, List, Optional, Union

import requests

# Module-level storage of last full response (for front-end retrieval if needed)
_LAST_RESPONSE: Dict[str, Any] = {}


def get_last_response() -> Dict[str, Any]:
    """Return the last raw response dict produced by ai_chat."""
    return _LAST_RESPONSE.copy()


def _detect_provider(base_url: Optional[str] = None) -> str:
    """Return provider name: 'deepseek', 'dashscope', or 'legacy'.

    Priority:
    1) Explicit AI_PROVIDER
    2) base_url hint
    3) API key env vars
    """
    prov = (os.environ.get('AI_PROVIDER') or '').strip().lower()
    if prov in {'deepseek', 'dashscope'}:
        return prov

    bu = (base_url or '').lower()
    if 'dashscope' in bu or 'aliyuncs.com' in bu:
        return 'dashscope'
    if 'deepseek' in bu:
        return 'deepseek'

    if os.environ.get('DASHSCOPE_API_KEY'):
        return 'dashscope'
    if os.environ.get('DEEPSEEK_API_KEY'):
        return 'deepseek'
    return 'legacy'


def _use_deepseek() -> bool:
    # Backwards-compatible wrapper for older callers.
    return _detect_provider() == 'deepseek'


def _normalize_openai_compatible_base(base_url: str) -> str:
    base = (base_url or '').strip().rstrip('/')
    if not base:
        return base
    if base.endswith('/v1'):
        return base
    return base + '/v1'


def _cap_max_tokens(provider: str, max_tokens: Optional[int]) -> Optional[int]:
    if max_tokens is None:
        return None
    try:
        mt = int(max_tokens)
    except Exception:
        return None

    if mt < 1:
        mt = 1

    # Provider-specific output-token caps (best-effort).
    # DeepSeek validates max_tokens and currently enforces <= 8192.
    provider_caps = {
        'deepseek': 8192,
        # DashScope compatible-mode models (e.g., qwen-plus) commonly enforce similar limits.
        # Keep conservative to avoid 400s; can be raised if you confirm higher limits.
        'dashscope': 8192,
    }
    cap = provider_caps.get((provider or '').lower())
    if cap is not None and mt > cap:
        return cap
    return mt


def _call_openai_compatible_via_requests(*,
                                        provider: str,
                                        api_key: str,
                                        base_url: str,
                                        model: str,
                                        messages: List[Dict[str, Any]],
                                        temperature: float,
                                        max_tokens: int,
                                        timeout_s: float = 600.0) -> Dict[str, Any]:
    base = _normalize_openai_compatible_base(base_url)
    url = f"{base}/chat/completions"
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }
    payload: Dict[str, Any] = {
        'model': model,
        'messages': messages,
        'temperature': temperature,
        'max_tokens': max_tokens,
    }
    r = requests.post(url, headers=headers, json=payload, timeout=timeout_s)
    try:
        data = r.json()
    except Exception:
        data = {'error': f'Non-JSON response (status={r.status_code})', 'text': r.text}

    if r.status_code >= 400:
        # Best-effort normalize to a consistent error envelope.
        return {
            'error': f"{provider} request failed (status={r.status_code})",
            'raw': data,
        }
    return data


def check_zhipu_model() -> tuple[bool, str]:
    """Probe available SDKs. Returns (ok, message)."""
    prov = _detect_provider()
    if prov in {'deepseek', 'dashscope'}:
        # We can use either the openai SDK or our requests fallback.
        try:
            import openai  # type: ignore
            return True, f'openai package available ({prov} compatible)'
        except Exception:
            return True, f'openai SDK missing; using requests fallback ({prov} compatible)'
    else:
        # try legacy zhipu adapter (best-effort)
        try:
            from reverie.backend_server.ai_providers import zhipu_api as legacy  # type: ignore
            if hasattr(legacy, 'check_zhipu_model'):
                return legacy.check_zhipu_model()
            return True, 'legacy zhipu adapter present'
        except Exception as e:
            return False, f'legacy zhipu adapter import failed: {e}'


import mimetypes
import base64

def prepare_multimodal_messages(messages: List[Dict[str, Any]], prompt_content: Optional[str] = None, files: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """
    Helper to prepare messages with multimodal content (files/images).
    Returns a new list of messages.
    """
    # Create a deep copy to avoid modifying the original list in place if needed, 
    # but for now we just create a new list with the last message modified.
    # Actually, let's just copy the list structure.
    new_messages = [m.copy() for m in messages]
    
    if not new_messages or new_messages[-1]['role'] != 'user':
        new_messages.append({'role': 'user', 'content': ''})
    
    last_msg = new_messages[-1]
    
    # Optimization: If no files are provided, prefer simple string content
    if not files and isinstance(last_msg['content'], str):
        if prompt_content:
            if last_msg['content']:
                last_msg['content'] += "\n" + prompt_content
            else:
                last_msg['content'] = prompt_content
        
        # Ensure content is not empty
        if not last_msg['content'] or (isinstance(last_msg['content'], str) and not last_msg['content'].strip()):
            last_msg['content'] = "..."
            
    else:
        # Convert existing string content to list if needed
        current_content = []
        if isinstance(last_msg['content'], str):
            if last_msg['content']:
                current_content.append({"type": "text", "text": last_msg['content']})
        elif isinstance(last_msg['content'], list):
            current_content = last_msg['content'] # Assume it's already a list of dicts
        
        # Add the new prompt content
        if prompt_content:
             current_content.append({"type": "text", "text": prompt_content})

        # Process Files
        if files:
            for file_path in files:
                mime_type, _ = mimetypes.guess_type(file_path)
                if not mime_type:
                    mime_type = 'application/octet-stream'
                
                if mime_type.startswith('image/'):
                    # Native Image Support (Vision)
                    try:
                        with open(file_path, "rb") as image_file:
                            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                            current_content.append({
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            })
                    except Exception as e:
                        current_content.append({"type": "text", "text": f"[Error reading image {file_path}: {e}]"})
                
                elif mime_type.startswith('audio/'):
                    # Audio Support
                    try:
                        # For now, we pass the file path or base64. 
                        # Many APIs prefer a URL or base64. Let's use base64 for consistency with images if small,
                        # but for audio/video, paths/URLs are often better. 
                        # However, to be generic for the Custom Generator, let's provide a structured object.
                        with open(file_path, "rb") as audio_file:
                            base64_audio = base64.b64encode(audio_file.read()).decode('utf-8')
                            current_content.append({
                                "type": "audio_url",
                                "audio_url": {
                                    "url": f"data:{mime_type};base64,{base64_audio}"
                                }
                            })
                    except Exception as e:
                        current_content.append({"type": "text", "text": f"[Error reading audio {file_path}: {e}]"})

                elif mime_type.startswith('video/'):
                    # Video Support
                    try:
                        # Video files can be large, so base64 might be risky. 
                        # But for consistency in this local simulation env:
                        with open(file_path, "rb") as video_file:
                            base64_video = base64.b64encode(video_file.read()).decode('utf-8')
                            current_content.append({
                                "type": "video_url",
                                "video_url": {
                                    "url": f"data:{mime_type};base64,{base64_video}"
                                }
                            })
                    except Exception as e:
                        current_content.append({"type": "text", "text": f"[Error reading video {file_path}: {e}]"})

                else:
                    # Document Parsing (PDF, Docx, Excel, Text)
                    try:
                        text_content = ""
                        ext = os.path.splitext(file_path)[1].lower()
                        
                        if ext == '.pdf':
                            try:
                                from pypdf import PdfReader
                                reader = PdfReader(file_path)
                                for page in reader.pages:
                                    text_content += page.extract_text() + "\n"
                            except ImportError:
                                text_content = "[Error: pypdf not installed]"
                        
                        elif ext in ['.docx', '.doc']:
                            try:
                                import docx
                                doc = docx.Document(file_path)
                                text_content = "\n".join([para.text for para in doc.paragraphs])
                            except ImportError:
                                text_content = "[Error: python-docx not installed]"
                                
                        elif ext in ['.xlsx', '.xls']:
                            try:
                                import openpyxl
                                wb = openpyxl.load_workbook(file_path, data_only=True)
                                for sheet in wb.sheetnames:
                                    text_content += f"Sheet: {sheet}\n"
                                    ws = wb[sheet]
                                    for row in ws.iter_rows(values_only=True):
                                        text_content += "\t".join([str(c) for c in row if c is not None]) + "\n"
                            except ImportError:
                                text_content = "[Error: openpyxl not installed]"
                        
                        else:
                            # Fallback to plain text
                            with open(file_path, "r", encoding="utf-8", errors='ignore') as f:
                                text_content = f.read()

                        current_content.append({"type": "text", "text": f"--- File: {os.path.basename(file_path)} ---\n{text_content}\n--- End of File ---"})
                    
                    except Exception as e:
                         current_content.append({"type": "text", "text": f"[Error reading file {file_path}: {e}]"})

        # Final Optimization: If content is a list but only contains text, flatten it back to string
        # This ensures compatibility with strict text-only models (like DashScope qwen-turbo) that reject list content.
        is_all_text = True
        combined_text = ""
        if isinstance(current_content, list):
            for item in current_content:
                if item.get('type') != 'text':
                    is_all_text = False
                    break
                combined_text += item.get('text', '') + "\n"
            
            if is_all_text and combined_text:
                last_msg['content'] = combined_text.strip()
            else:
                last_msg['content'] = current_content
        else:
            last_msg['content'] = current_content
        
    return new_messages

def extract_content_from_raw(raw_data: Any) -> Optional[str]:
    """
    Helper to extract content (text/images) from a raw response object/dict.
    Returns a JSON string of multimodal content if images found, or plain text, or None.
    """
    content = None
    try:
        # 1. Try DashScope specific structure: output.choices[0].message.content
        if 'output' in raw_data and 'choices' in raw_data['output']:
            ds_msg = raw_data['output']['choices'][0]['message']
            ds_content = ds_msg.get('content')
            
            if isinstance(ds_content, list):
                # Convert DashScope image format to OpenAI format
                normalized_content = []
                for item in ds_content:
                    if isinstance(item, dict):
                        if 'image' in item:
                            normalized_content.append({
                                "type": "image_url",
                                "image_url": {"url": item['image']}
                            })
                        elif 'text' in item:
                            normalized_content.append({
                                "type": "text",
                                "text": item['text']
                            })
                if normalized_content:
                    content = json.dumps(normalized_content)
            elif isinstance(ds_content, str):
                content = ds_content
    except Exception:
        pass

    if content:
        return content

    # 2. Recursive Image Search Fallback
    try:
        def find_images(obj):
            found = []
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if k == 'image' and isinstance(v, str) and v.startswith('http'):
                        found.append(v)
                    else:
                        found.extend(find_images(v))
            elif isinstance(obj, list):
                for item in obj:
                    found.extend(find_images(item))
            return found

        images = find_images(raw_data)
        if images:
            normalized_content = []
            for img_url in images:
                normalized_content.append({
                    "type": "image_url",
                    "image_url": {"url": img_url}
                })
            content = json.dumps(normalized_content)
    except Exception:
        pass
        
    return content

def ai_chat(messages_or_prompt: Union[str, List[Dict[str, str]]], *,
            prompt_content: Optional[str] = None,
            files: Optional[List[str]] = None,
            model: Optional[str] = None,
            api_key: Optional[str] = None,
            base_url: Optional[str] = None,
            temperature: Optional[float] = None,
            max_tokens: Optional[int] = None,
            timeout_s: Optional[float] = None) -> Dict[str, Any]:
    """Return a dict with content, reasoning_content (if available), and raw response.

    messages_or_prompt: either a prompt string (will be wrapped as a single user message)
    or a list of messages in OpenAI chat format.
    prompt_content: The main text prompt for the current step (if messages_or_prompt is a list, this is appended to the user message).
    files: List of absolute file paths to attach.
    """
    global _LAST_RESPONSE
    provider = _detect_provider(base_url)
    capped_max_tokens = _cap_max_tokens(provider, max_tokens)
    req_timeout = 600.0
    if timeout_s is not None:
        try:
            req_timeout = float(timeout_s)
        except Exception:
            req_timeout = 600.0
        if req_timeout <= 0:
            req_timeout = 600.0
    
    # Normalize messages
    if isinstance(messages_or_prompt, str):
        messages = [{'role': 'user', 'content': messages_or_prompt}]
    else:
        messages = messages_or_prompt

    # Use helper to prepare messages
    messages = prepare_multimodal_messages(messages, prompt_content, files)

    if provider in {'deepseek', 'dashscope'}:
        # OpenAI-compatible path (DeepSeek / DashScope compatible-mode)
        # Support both the new `from openai import OpenAI` client and the older
        # legacy `import openai` module. Try new API first, fall back to legacy.
        if provider == 'dashscope':
            ds_key = api_key or os.environ.get('DASHSCOPE_API_KEY') or os.environ.get('DEEPSEEK_API_KEY')
            ds_base = base_url or os.environ.get('DASHSCOPE_BASE_URL') or 'https://dashscope.aliyuncs.com/compatible-mode/v1'
            ds_model = model or os.environ.get('DASHSCOPE_DEFAULT_MODEL') or 'qwen-plus'
        else:
            ds_key = api_key or os.environ.get('DEEPSEEK_API_KEY')
            ds_base = base_url or os.environ.get('DEEPSEEK_BASE_URL') or 'https://api.deepseek.com'
            ds_model = model or os.environ.get('DEEPSEEK_DEFAULT_MODEL') or 'deepseek-reasoner'

        if not ds_key:
            _LAST_RESPONSE = {'error': f"Missing API key for provider={provider}. Set {'DASHSCOPE_API_KEY' if provider=='dashscope' else 'DEEPSEEK_API_KEY'} or pass api_key."}
            return _LAST_RESPONSE

        # Attempt to get new-style client
        client = None
        client_type = None
        
        # DEBUG: Write payload to file
        try:
            with open("debug_payload.json", "w", encoding="utf-8") as f:
                json.dump({
                    "url": ds_base,
                    "model": ds_model,
                    "messages": messages
                }, f, ensure_ascii=False, indent=2)
        except:
            pass

        try:
            from openai import OpenAI  # type: ignore
            client = OpenAI(api_key=ds_key, base_url=ds_base)
            client_type = 'new'
        except Exception:
            try:
                import openai  # type: ignore
                # configure legacy openai module
                try:
                    openai.api_key = ds_key
                except Exception:
                    pass
                try:
                    openai.api_base = _normalize_openai_compatible_base(ds_base)
                except Exception:
                    pass
                client = openai
                client_type = 'legacy'
            except Exception:
                client = None
                client_type = 'requests'

        try:
            if client_type == 'new':
                resp = client.chat.completions.create(
                    model=ds_model,
                    messages=messages,
                    temperature=(temperature if temperature is not None else 0.0),
                    max_tokens=(capped_max_tokens if capped_max_tokens is not None else 768),
                    timeout=req_timeout,
                )
                try:
                    choice = resp.choices[0].message
                    content = getattr(choice, 'content', None)
                    reasoning = getattr(choice, 'reasoning_content', None)
                except Exception:
                    content = None
                    reasoning = None
            elif client_type == 'legacy':
                try:
                    resp = client.ChatCompletion.create(
                        model=ds_model,
                        messages=messages,
                        temperature=(temperature if temperature is not None else 0.0),
                        max_tokens=(capped_max_tokens if capped_max_tokens is not None else 512),
                        request_timeout=req_timeout,
                    )
                except Exception:
                    resp = client.ChatCompletion.create(
                        model=ds_model,
                        messages=messages,
                        temperature=(temperature if temperature is not None else 0.0),
                        max_tokens=(capped_max_tokens if capped_max_tokens is not None else 512),
                        request_timeout=req_timeout,
                    )
                try:
                    content = resp['choices'][0]['message'].get('content')
                except Exception:
                    try:
                        content = resp.choices[0].message['content']
                    except Exception:
                        content = None
                reasoning = None
            else:
                resp = _call_openai_compatible_via_requests(
                    provider=provider,
                    api_key=ds_key,
                    base_url=ds_base,
                    model=ds_model,
                    messages=messages,
                    temperature=(temperature if temperature is not None else 0.0),
                    max_tokens=(capped_max_tokens if capped_max_tokens is not None else 768),
                    timeout_s=req_timeout,
                )
                if isinstance(resp, dict) and resp.get('error') and resp.get('raw'):
                    _LAST_RESPONSE = {'error': resp.get('error'), 'raw': resp.get('raw')}
                    return _LAST_RESPONSE
                try:
                    content = (resp.get('choices') or [None])[0].get('message', {}).get('content')
                except Exception:
                    content = None
                try:
                    reasoning = (resp.get('choices') or [None])[0].get('message', {}).get('reasoning_content')
                except Exception:
                    reasoning = None

            # --- DashScope / Wanx Image Generation Fix ---
            # Some DashScope models return a native format structure even via compatible endpoints,
            # or if the user is using a specific model that returns 'output' -> 'choices'.
            # Structure: {"output": {"choices": [{"message": {"content": [{"image": "..."}]}}]}}
            if not content and not reasoning:
                try:
                    # Try to access as dict or object
                    if isinstance(resp, dict):
                        raw_data = resp
                    elif hasattr(resp, 'model_dump'):
                        raw_data = resp.model_dump()
                    elif hasattr(resp, 'to_dict_recursive'):
                        raw_data = resp.to_dict_recursive()
                    elif hasattr(resp, 'to_dict'):
                        raw_data = resp.to_dict()
                    else:
                        raw_data = resp.__dict__
                    
                    # Use shared extraction logic
                    extracted = extract_content_from_raw(raw_data)
                    if extracted:
                        content = extracted
                        
                except Exception as e:
                    print(f"Error parsing DashScope response: {e}")

            # --- Smart Fallback: AI-driven Parsing ---
            # If we still have no content, ask the AI to parse its own raw response.
            # This handles unknown/proprietary formats dynamically.
            if not content and not reasoning and 'raw_data' in locals():
                try:
                    print("Standard parsing failed. Attempting AI-driven parsing...")
                    parsing_prompt = [
                        {"role": "system", "content": "You are a JSON data extraction assistant. Your job is to extract the main 'content' (text) and any 'image' URLs from the provided raw JSON response."},
                        {"role": "user", "content": f"""
I have a raw JSON response from an AI API, but I cannot parse it with standard rules.
Please extract the text content and any image URLs.

Raw JSON Data:
{json.dumps(raw_data, default=str)[:3000]} 

Instructions:
1. Return ONLY a valid JSON List.
2. Use this format: [{{"type": "text", "text": "..."}}, {{"type": "image_url", "image_url": {{"url": "..."}}}}]
3. Do not include markdown formatting (```json).
4. If the JSON contains an error message, extract that as text.
"""}
                    ]
                    
                    # Reuse the existing client to make the parsing call
                    # We use a low temperature for deterministic parsing
                    if client_type == 'new':
                        parse_resp = client.chat.completions.create(
                            model=ds_model, 
                            messages=parsing_prompt,
                            temperature=0.0,
                            max_tokens=1500
                        )
                        parsed_text = parse_resp.choices[0].message.content
                    else:
                        # legacy
                        parse_resp = client.ChatCompletion.create(
                            model=ds_model, 
                            messages=parsing_prompt,
                            temperature=0.0,
                            max_tokens=1500
                        )
                        parsed_text = parse_resp.choices[0].message['content']
                    
                    # Clean up potential markdown
                    parsed_text = parsed_text.strip()
                    if parsed_text.startswith("```json"):
                        parsed_text = parsed_text[7:]
                    if parsed_text.startswith("```"):
                        parsed_text = parsed_text[3:]
                    if parsed_text.endswith("```"):
                        parsed_text = parsed_text[:-3]
                    
                    # Validate JSON
                    json.loads(parsed_text) 
                    content = parsed_text.strip()
                    print("AI-driven parsing successful.")
                    
                except Exception as e:
                    print(f"AI-driven parsing failed: {e}")
            # ---------------------------------------------

            _LAST_RESPONSE = {'provider': provider, 'content': content, 'reasoning_content': reasoning, 'raw': resp}
            return _LAST_RESPONSE
        except Exception as e:
            error_msg = f"Request failed: {e}"
            if hasattr(e, 'response') and hasattr(e.response, 'text'):
                 error_msg += f" | Server Response: {e.response.text}"
            elif hasattr(e, 'body'):
                 error_msg += f" | Body: {e.body}"
            
            # --- Auto-Retry for Text-Only Models ---
            # If the error indicates that the model rejected 'image_url' (multimodal content),
            # we automatically strip the images and retry with text only.
            # Common error: "unknown variant `image_url`, expected `text`"
            if "image_url" in str(e) and "expected" in str(e) and "text" in str(e):
                print("Multimodal request failed (model likely text-only). Retrying with text only...")
                try:
                    # Strip images from messages
                    text_only_messages = []
                    for m in messages:
                        new_m = m.copy()
                        if isinstance(new_m.get('content'), list):
                            # Filter out image_url, keep text
                            text_parts = [p.get('text', '') for p in new_m['content'] if p.get('type') == 'text']
                            # Add a note about removed images
                            if any(p.get('type') == 'image_url' for p in new_m['content']):
                                text_parts.append("\n[System Note: An image was attached but removed because this model does not support vision.]")
                            new_m['content'] = "\n".join(text_parts)
                        text_only_messages.append(new_m)
                    
                    # Retry call
                    if client_type == 'new':
                        resp = client.chat.completions.create(model=ds_model, messages=text_only_messages,
                                                              temperature=(temperature if temperature is not None else 0.0),
                                                              max_tokens=(max_tokens if max_tokens is not None else 768))
                        choice = resp.choices[0].message
                        content = getattr(choice, 'content', None)
                        reasoning = getattr(choice, 'reasoning_content', None)
                    else:
                        resp = client.ChatCompletion.create(model=ds_model, messages=text_only_messages, temperature=(temperature if temperature is not None else 0.0), max_tokens=(max_tokens if max_tokens is not None else 768))
                        try:
                            content = resp['choices'][0]['message'].get('content')
                        except:
                            content = resp.choices[0].message['content']
                        reasoning = None
                        
                    _LAST_RESPONSE = {'provider': 'deepseek', 'content': content, 'reasoning_content': reasoning, 'raw': resp}
                    return _LAST_RESPONSE
                except Exception as retry_e:
                    print(f"Retry failed: {retry_e}")
                    error_msg += f" | Retry Failed: {retry_e}"
            # ---------------------------------------
            
            print(f"ERROR: {error_msg}")
            try:
                with open("debug_error.log", "w", encoding="utf-8") as f:
                    f.write(error_msg)
            except:
                pass
            _LAST_RESPONSE = {'error': error_msg}
            return _LAST_RESPONSE
    else:
        # Try legacy zhipu adapter if available
        try:
            from reverie.backend_server.ai_providers import zhipu_api as legacy  # type: ignore
            if hasattr(legacy, 'ai_chat'):
                out = legacy.ai_chat(messages, model=model, api_key=api_key, base_url=base_url,
                                     temperature=temperature, max_tokens=max_tokens)
                _LAST_RESPONSE = {'provider': 'zhipu_legacy', 'content': out.get('content'), 'reasoning_content': out.get('reasoning_content'), 'raw': out.get('raw')}
                return _LAST_RESPONSE
            else:
                # legacy adapter may only provide zhipu_chat returning string
                if isinstance(messages_or_prompt, str):
                    prompt = messages_or_prompt
                else:
                    # join messages into a single prompt
                    prompt = '\n'.join([m.get('content','') for m in messages])
                content = legacy.zhipu_chat(prompt, model=model, api_key=api_key)
                _LAST_RESPONSE = {'provider': 'zhipu_legacy', 'content': content, 'reasoning_content': None, 'raw': content}
                return _LAST_RESPONSE
        except Exception as e:
            _LAST_RESPONSE = {'error': f'no adapter available: {e}'}
            return _LAST_RESPONSE


def zhipu_chat(prompt: str, *, model: Optional[str] = None, api_key: Optional[str] = None,
               temperature: Optional[float] = None, max_tokens: Optional[int] = None) -> str:
    """Compatibility wrapper: returns content string (legacy behavior).
    For DeepSeek it will call ai_chat and return the content; full response
    is available via get_last_response()."""
    res = ai_chat(prompt, model=model, api_key=api_key, base_url=None, temperature=temperature, max_tokens=max_tokens)
    if res is None:
        return ''
    # Prefer the regular 'content' field by default; fall back to reasoning_content if content missing
    if 'content' in res and res.get('content') is not None:
        return res.get('content')
    if res.get('reasoning_content'):
        return res.get('reasoning_content')
    if 'raw' in res and isinstance(res['raw'], str):
        return res['raw']
    # fallback - stringify
    try:
        return str(res)
    except Exception:
        return ''


def zhipu_embedding(text: str, *, model: Optional[str] = None, api_key: Optional[str] = None) -> Optional[List[float]]:
    """Try DeepSeek/OpenAI-compatible embedding; fallback to legacy adapter or None."""
    if _use_deepseek():
        # similar dual-client support for embeddings
        ds_key = api_key or os.environ.get('DEEPSEEK_API_KEY')
        ds_base = os.environ.get('DEEPSEEK_BASE_URL') or 'https://api.deepseek.com'
        ds_model = model or os.environ.get('DEEPSEEK_DEFAULT_MODEL') or 'text-embedding-ada-002'

        try:
            from openai import OpenAI  # type: ignore
            client_new = OpenAI(api_key=ds_key, base_url=ds_base)
            try:
                r = client_new.embeddings.create(model=ds_model, input=text)
                emb = None
                try:
                    emb = r.data[0].embedding
                except Exception:
                    try:
                        emb = r['data'][0]['embedding']
                    except Exception:
                        emb = None
                return emb
            except Exception:
                return None
        except Exception:
            try:
                import openai  # type: ignore
                try:
                    openai.api_key = ds_key
                except Exception:
                    pass
                try:
                    if not ds_base.endswith('/v1'):
                        openai.api_base = ds_base.rstrip('/') + '/v1'
                    else:
                        openai.api_base = ds_base
                except Exception:
                    pass
                try:
                    r = openai.Embedding.create(model=ds_model, input=text)
                    try:
                        return r['data'][0]['embedding']
                    except Exception:
                        try:
                            return r.data[0].embedding
                        except Exception:
                            return None
                except Exception:
                    return None
            except Exception:
                return None
    else:
        try:
            from reverie.backend_server.ai_providers import zhipu_api as legacy  # type: ignore
            if hasattr(legacy, 'zhipu_embedding'):
                return legacy.zhipu_embedding(text, model=model, api_key=api_key)
        except Exception:
            pass
    return None
